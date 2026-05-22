import json

import csv
from io import BytesIO, StringIO
try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - dependency is declared in requirements.txt
    Workbook = load_workbook = None
from collections import defaultdict
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from jobradar.models import ApplicationNote, FollowUp, JobEvaluation, JobLead, UserProfile

SCHEMA_VERSION = 1
APP_NAME = 'dachapply'

JOB_FIELDS = ['company','title','location','url','source','raw_description','submitted_by','submitter_reason','salary_info','language_requirements','work_mode','status','status_date','interview_stage','interview_total','last_update_date','feedback_due_date']
EVALUATION_FIELDS = ['fit_score','priority','recommendation','summary','main_match_reasons','main_gaps','required_skills','nice_to_have_skills','matched_skills','missing_skills','cv_adjustment_notes','interview_prep_notes','risk_notes','next_action','structured_json_raw']
NOTE_FIELDS = ['note','note_type']
FOLLOWUP_FIELDS = ['follow_up_date','reason','completed']
PROFILE_FIELDS = []

# InviteCode is intentionally excluded: ownership is ambiguous and codes can act as
# access credentials/secrets. Django auth/session/admin/permission/token models are
# also excluded because exports must never contain passwords, sessions, tokens, logs,
# permissions, or secrets.

def _iso(value):
    return value.isoformat() if hasattr(value, 'isoformat') else value

def _clean_record(obj, fields, extra=None):
    data = {'id': obj.id}
    for field in fields:
        data[field] = _iso(getattr(obj, field))
    if extra:
        data.update(extra)
    return data

def owned_jobs(user):
    return JobLead.objects.filter(Q(created_by=user) | Q(submitted_for=user)).distinct()

def build_user_export(user):
    jobs = owned_jobs(user).prefetch_related('evaluations', 'notes', 'followups')
    job_ids = list(jobs.values_list('id', flat=True))
    profile = getattr(user, 'jobradar_profile', None)
    return {
        'schema_version': SCHEMA_VERSION,
        'exported_at': timezone.now().isoformat(),
        'app': APP_NAME,
        'user': {'id': user.id, 'username': user.get_username(), 'email': getattr(user, 'email', '') or ''},
        'data': {
            'profile': [_clean_record(profile, PROFILE_FIELDS)] if profile else [],
            'jobs': [_clean_record(j, JOB_FIELDS, {'created_by_username': j.created_by.username if j.created_by else '', 'submitted_for_username': j.submitted_for.username if j.submitted_for else ''}) for j in jobs],
            'evaluations': [_clean_record(e, EVALUATION_FIELDS, {'job': e.job_id}) for e in JobEvaluation.objects.filter(job_id__in=job_ids)],
            'notes': [_clean_record(n, NOTE_FIELDS, {'job': n.job_id}) for n in ApplicationNote.objects.filter(job_id__in=job_ids)],
            'followups': [_clean_record(f, FOLLOWUP_FIELDS, {'job': f.job_id}) for f in FollowUp.objects.filter(job_id__in=job_ids)],
        },
    }


def _flatten_for_table(value):
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return _iso(value)

def _unflatten_value(value):
    if value == '':
        return ''
    if isinstance(value, str) and value[:1] in ('[', '{'):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value

def export_user_data_csv(user):
    export = build_user_export(user)
    rows = export['data']['jobs']
    output = StringIO()
    fields = ['id'] + JOB_FIELDS
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction='ignore')
    writer.writeheader()
    for row in rows:
        writer.writerow({field: _flatten_for_table(row.get(field)) for field in fields})
    return output.getvalue()

def export_user_data_xlsx(user):
    if Workbook is None:
        raise RuntimeError('openpyxl is required for XLSX export')
    export = build_user_export(user)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    sheet_fields = {
        'profile': ['id'],
        'jobs': ['id'] + JOB_FIELDS,
        'evaluations': ['id', 'job'] + EVALUATION_FIELDS,
        'notes': ['id', 'job'] + NOTE_FIELDS,
        'followups': ['id', 'job'] + FOLLOWUP_FIELDS,
    }
    for name, fields in sheet_fields.items():
        sheet = workbook.create_sheet(name)
        sheet.append(fields)
        for row in export['data'].get(name, []):
            sheet.append([_flatten_for_table(row.get(field)) for field in fields])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

def _payload_from_csv_file(file_obj):
    text = file_obj.read().decode('utf-8-sig')
    rows = []
    for row in csv.DictReader(StringIO(text)):
        rows.append({k: _unflatten_value(v) for k, v in row.items() if k})
    return {'schema_version': SCHEMA_VERSION, 'app': APP_NAME, 'data': {'jobs': rows, 'evaluations': [], 'notes': [], 'followups': [], 'profile': []}}

def _payload_from_xlsx_file(file_obj):
    if load_workbook is None:
        raise ValueError('XLSX import is not available')
    workbook = load_workbook(filename=BytesIO(file_obj.read()), data_only=True)
    data = {'profile': [], 'jobs': [], 'evaluations': [], 'notes': [], 'followups': []}
    for sheet_name in data.keys():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else '' for h in rows[0]]
        for values in rows[1:]:
            row = {headers[i]: _unflatten_value('' if values[i] is None else str(values[i])) for i in range(min(len(headers), len(values))) if headers[i]}
            if any(v not in ('', None) for v in row.values()):
                data[sheet_name].append(row)
    return {'schema_version': SCHEMA_VERSION, 'app': APP_NAME, 'data': data}

def parse_import_payload(request):
    def with_options(payload):
        if hasattr(request, 'data'):
            if request.data.get('duplicate_strategy'):
                payload['duplicate_strategy'] = request.data.get('duplicate_strategy')
            if request.data.get('duplicate_actions'):
                try:
                    payload['duplicate_actions'] = json.loads(request.data.get('duplicate_actions')) if isinstance(request.data.get('duplicate_actions'), str) else request.data.get('duplicate_actions')
                except json.JSONDecodeError:
                    raise ValueError('duplicate_actions must be valid JSON')
        return payload
    try:
        upload = request.FILES.get('file') if hasattr(request, 'FILES') else None
        if upload:
            name = upload.name.lower()
            if name.endswith('.csv'):
                return with_options(_payload_from_csv_file(upload))
            if name.endswith('.xlsx'):
                return with_options(_payload_from_xlsx_file(upload))
            if not name.endswith('.json'):
                raise ValueError('Unsupported import file type. Use JSON, CSV, or XLSX.')
            parsed = json.loads(upload.read().decode('utf-8-sig'))
        else:
            parsed = request.data
    except json.JSONDecodeError as exc:
        raise ValueError(f'Invalid JSON: {exc.msg}')
    except Exception as exc:
        raise ValueError(str(exc))
    if isinstance(parsed, dict) and isinstance(parsed.get('json'), str):
        try:
            parsed = json.loads(parsed['json'])
        except json.JSONDecodeError as exc:
            raise ValueError(f'Invalid JSON: {exc.msg}')
    return with_options(parsed)

def _parse_value(field, value):
    if value in ('', None):
        return None if field.endswith('_date') else value
    if field.endswith('_date'):
        return parse_date(value) if isinstance(value, str) else value
    return value

def _assign_fields(obj, fields, data):
    changed = False
    for field in fields:
        if field not in data:
            continue
        value = _parse_value(field, data[field])
        if getattr(obj, field) != value:
            setattr(obj, field, value)
            changed = True
    return changed

def _find_owned_by_import_id(model, old_id, user, job_map=None):
    if not old_id:
        return None
    if model is JobLead:
        return owned_jobs(user).filter(id=old_id).first()
    if model is JobEvaluation:
        return model.objects.filter(id=old_id, job__in=owned_jobs(user)).first()
    if model is ApplicationNote:
        return model.objects.filter(id=old_id, job__in=owned_jobs(user)).first()
    if model is FollowUp:
        return model.objects.filter(id=old_id, job__in=owned_jobs(user)).first()
    return None

def _import_action(payload, index):
    for action in payload.get('duplicate_actions') or []:
        if action.get('index') == index:
            return action.get('action') or 'skip'
    return payload.get('duplicate_strategy')

def _changed_fields(obj, fields, data):
    changed = []
    for field in fields:
        if field not in data:
            continue
        if getattr(obj, field) != _parse_value(field, data[field]):
            changed.append(field)
    return changed

def _job_conflicts(user, payload):
    conflicts = []
    for i, item in enumerate(payload.get('data', {}).get('jobs', [])):
        if not isinstance(item, dict) or _import_action(payload, i):
            continue
        owned = owned_jobs(user)
        by_id = _find_owned_by_import_id(JobLead, item.get('id'), user)
        if by_id:
            changed = _changed_fields(by_id, JOB_FIELDS, item)
            if changed:
                conflicts.append({'index': i, 'kind': 'update', 'incoming': {'company': item.get('company'), 'title': item.get('title'), 'url': item.get('url')}, 'existing_jobs': [{'id': by_id.id, 'company': by_id.company, 'title': by_id.title, 'url': by_id.url, 'status': by_id.status}], 'changed_fields': changed})
            continue
        if item.get('url'):
            existing = list(owned.filter(url=item.get('url')).values('id', 'company', 'title', 'url', 'status')[:10])
            if existing:
                conflicts.append({'index': i, 'kind': 'duplicate_url', 'url': item.get('url'), 'incoming': {'company': item.get('company') or 'Unknown company', 'title': item.get('title') or 'Untitled role', 'url': item.get('url')}, 'existing_jobs': existing})
    return conflicts

def import_user_export(user, payload):
    if not isinstance(payload, dict):
        return {'created': {}, 'updated': {}, 'skipped': {}, 'errors': ['Import payload must be a JSON object']}
    if payload.get('schema_version') != SCHEMA_VERSION:
        return {'created': {}, 'updated': {}, 'skipped': {}, 'errors': ['Unsupported schema_version']}
    if payload.get('app') != APP_NAME or not isinstance(payload.get('data'), dict):
        return {'created': {}, 'updated': {}, 'skipped': {}, 'errors': ['Invalid export structure']}

    conflicts = _job_conflicts(user, payload)
    if conflicts:
        return {'type': 'import_conflicts', 'message': 'Some imported records conflict with existing data. Choose override, duplicate, skip, or abort.', 'conflicts': conflicts, 'created': {}, 'updated': {}, 'skipped': {}, 'errors': []}

    summary = {'created': defaultdict(int), 'updated': defaultdict(int), 'skipped': defaultdict(int), 'errors': []}
    data = payload['data']
    job_map = {}

    with transaction.atomic():
        for item in data.get('profile', []):
            profile, created = UserProfile.objects.get_or_create(user=user)
            summary['created' if created else 'skipped']['profile'] += 1

        for i, item in enumerate(data.get('jobs', [])):
            if not isinstance(item, dict):
                summary['errors'].append('Skipped invalid job record')
                continue
            if not item.get('submitted_by') and item.get('created_by_username'):
                item = {**item, 'submitted_by': item.get('created_by_username')}
            action = _import_action(payload, i)
            obj = _find_owned_by_import_id(JobLead, item.get('id'), user)
            if not obj and item.get('url') and action == 'override':
                obj = owned_jobs(user).filter(url=item.get('url')).first()
            if action == 'skip':
                summary['skipped']['jobs'] += 1
                continue
            if obj and action == 'duplicate':
                obj = None
            if obj:
                changed = _assign_fields(obj, JOB_FIELDS, item)
                obj.created_by = user
                obj.submitted_for = None
                if changed:
                    obj.save()
                    summary['updated']['jobs'] += 1
                else:
                    obj.save(update_fields=['created_by','submitted_for'])
                    summary['skipped']['jobs'] += 1
            else:
                obj = JobLead(created_by=user, submitted_for=None)
                _assign_fields(obj, JOB_FIELDS, item)
                obj.company = obj.company or 'Unknown company'
                obj.title = obj.title or 'Untitled role'
                obj.save()
                summary['created']['jobs'] += 1
            if item.get('id'):
                job_map[item['id']] = obj

        for resource, model, fields in [('evaluations', JobEvaluation, EVALUATION_FIELDS), ('notes', ApplicationNote, NOTE_FIELDS), ('followups', FollowUp, FOLLOWUP_FIELDS)]:
            for item in data.get(resource, []):
                if not isinstance(item, dict):
                    summary['errors'].append(f'Skipped invalid {resource} record')
                    continue
                job = job_map.get(item.get('job')) or owned_jobs(user).filter(id=item.get('job')).first()
                if not job:
                    summary['skipped'][resource] += 1
                    continue
                obj = _find_owned_by_import_id(model, item.get('id'), user)
                if obj:
                    changed = _assign_fields(obj, fields, item)
                    if resource == 'notes':
                        obj.created_by = user
                    obj.job = job
                    if changed:
                        obj.save()
                        summary['updated'][resource] += 1
                    else:
                        obj.save()
                        summary['skipped'][resource] += 1
                else:
                    kwargs = {'job': job}
                    if resource == 'notes':
                        kwargs['created_by'] = user
                    obj = model(**kwargs)
                    _assign_fields(obj, fields, item)
                    obj.save()
                    summary['created'][resource] += 1

    return {k: dict(v) if hasattr(v, 'items') else v for k, v in summary.items()}
